from tkinter import*
from tkinter import ttk
from tkcalendar import DateEntry
import tkinter.messagebox as tmsg
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
root=Tk()
root.geometry("800x850")
root.title("Admission Form-->>")
Label(root,text="Please read & fill the form carefully",font="lucida 25 bold",fg="red").grid(row=0,column=3,pady=10)
name=Label(root,text="Enter Your Name",font="lucida 12 bold")
fname=Label(root,text="Father's Name",font="lucida 12 bold")
dob=Label(root,text="Date Of Birth",font="lucida 12 bold")
dept=Label(root,text="Depertment",font="lucida 12 bold")
gender=Label(root,text="Gender",font="lucida 12 bold")
phone=Label(root,text="Enter PH No",font="lucida 12 bold")
Email=Label(root,text="Email id",font="lucida 12 bold")
Aadhar=Label(root,text="Aadhar No",font="lucida 12 bold")
Address=Label(root,text="Permanent Address",font="lucida 12 bold")
Address2=Label(root,text="Current Address",font="lucida 12 bold")
xmarks=Label(root,text="10th Marks",font="lucida 12 bold")
xiimarks=Label(root,text="12th Marks",font="lucida 12 bold")


name.grid(row=1,column=2,pady=5)
fname.grid(row=2,column=2,pady=5)
dob.grid(row=3,column=2,pady=5)
dept.grid(row=4,column=2,pady=5)
gender.grid(row=5,column=2,pady=5)
phone.grid(row=8,column=2,pady=5)
Email.grid(row=9,column=2,pady=5)
Aadhar.grid(row=10,column=2,pady=5)
Address.grid(row=11,column=2,pady=5)
Address2.grid(row=12,column=2,pady=5)
xmarks.grid(row=13,column=2,pady=5)
xiimarks.grid(row=14,column=2,pady=5)
namevar=StringVar()
fnamevar=StringVar()
Gendervar=StringVar()
phonevar=StringVar()
Emailvar=StringVar()
Aadharvar=StringVar()
Addressvar=StringVar()
Address2var=StringVar()
xmarksvar=StringVar()
xiimarksvar=StringVar()
var=IntVar()
Gendervar.set("0")
deptbox=ttk.Combobox(root,values=["CSE","IT","ECE","AIML","EE"])
deptbox.grid(row=4,column=3,padx=10,pady=7)
deptbox.set("Choose Depertment")

dobcal=DateEntry(root, selectmode='day',date_pattern='dd-mm-yyyy')
dobcal.grid(row=3,column=3,padx=10,pady=10)

radio=Radiobutton(root,text="MALE",variable=Gendervar,value="MALE",anchor="w").grid(row=5,column=3)
radio=Radiobutton(root,text="FEMALE",variable=Gendervar,value="FEMALE",anchor="w").grid(row=6,column=3)
radio=Radiobutton(root,text="OTHER",variable=Gendervar,value="OTHER",anchor="w").grid(row=7,column=3)

name1=Entry(root,textvar=namevar,font="lucida 14 bold",bg="pink").grid(row=1,column=3,pady=5)
fname1=Entry(root,textvar=fnamevar,font="lucida 14 bold",bg="pink").grid(row=2,column=3,pady=5)
phone1=Entry(root,textvar=phonevar,font="lucida 14 bold",bg="pink").grid(row=8,column=3,pady=5)
Email1=Entry(root,textvar=Emailvar,font="lucida 14 bold",bg="pink").grid(row=9,column=3,pady=5)
Aadhar1=Entry(root,textvar=Aadharvar,font="lucida 14 bold",bg="pink").grid(row=10,column=3,pady=5)
Address1=Entry(root,textvar=Addressvar,font="lucida 14 bold",bg="pink").grid(row=11,column=3,pady=5)
Address21=Entry(root,textvar=Address2var,font="lucida 14 bold",bg="pink").grid(row=12,column=3,pady=5)
xmarks1=Entry(root,textvar=xmarksvar,font="lucida 14 bold",bg="pink").grid(row=13,column=3,pady=5)
xiimarks1=Entry(root,textvar=xiimarksvar,font="lucida 14 bold",bg="pink").grid(row=14,column=3,pady=5)
Checkbutton(root,variable=var,text="All the details provided are correct",bg="blue",pady=6,padx=10).grid(row=16,column=3)

def getvals():
    print("SUBMIT IS DONE\n")
    print(f"{namevar.get(),fnamevar.get(),dobcal.get(),deptbox.get(),Gendervar.get(),phonevar.get(),Emailvar.get(),Aadharvar.get(),Addressvar.get(),Address2var.get(),xmarksvar.get(),xiimarksvar.get()}")
    c=var.get()
    if c==0:
        tmsg.showwarning("Warning","click on the cheackbox before submit")
    elif c==1:
        file_path="form_details.xlsx"
        if not os.path.exists(file_path):
            wb=Workbook()
            ws=wb.active
            headers=["Name","Father's Name","Date Of Birth","Depertment","Gender",
            "Enter PH No","Email id","Aadhar No","Permanent Addre","Current Address",
            "10th Marks","12th Marks"]
            ws.append(headers)
            wb.save(file_path)
        from openpyxl import load_workbook
        wb=load_workbook(file_path)
        ws=wb.active
        new_row=[namevar.get(),fnamevar.get(),dobcal.get(),deptbox.get(),Gendervar.get(),phonevar.get(),Emailvar.get(),Aadharvar.get(),Addressvar.get(),Address2var.get(),xmarksvar.get(),xiimarksvar.get()]
        ws.append(new_row)
        wb.save(file_path)
        tmsg.showinfo("Submit","You have succesfully submitted the form")
Button(text="Submit",command=getvals,bg="grey").grid(column=3,pady=20)
root.mainloop()